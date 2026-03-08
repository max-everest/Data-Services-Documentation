import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import re

INPUT_FOLDER = "legacy_runbooks"
OUTPUT_FOLDER = "content/runbooks/drafts"

Path(OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)


SECTION_ORDER = [
    "Overview",
    "Quick Triage",
    "Diagnostics",
    "Resolution",
    "Escalation",
    "Communication"
]


def slugify(name):
    name = name.lower()
    name = re.sub(r"[^\w\s-]", "", name)
    name = re.sub(r"\s+", "-", name)
    return name


def title_case(name):
    name = name.replace("-", " ")
    return name.title()


for file in Path(INPUT_FOLDER).glob("*.xlsx"):

    wb = load_workbook(file)

    runbook_name = title_case(file.stem)

    md = []
    md.append(f"# {runbook_name}")
    md.append("")

    # Create lookup of sheets
    sheets = {name: wb[name] for name in wb.sheetnames}

    for section in SECTION_ORDER:

        if section not in sheets:
            continue

        ws = sheets[section]

        md.append(f"## {section}")
        md.append("")

        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        rows = [r for r in rows if any(c is not None for c in r)]

        if not rows:
            md.append("")
            continue

        df = pd.DataFrame(rows)
        df = df.dropna(axis=1, how="all")

        if df.shape[1] > 1:

            header = df.iloc[0]
            table = df[1:]

            md.append("| " + " | ".join(str(h) for h in header) + " |")
            md.append("|" + " --- |" * len(header))

            for _, row in table.iterrows():
                md.append("| " + " | ".join("" if pd.isna(x) else str(x) for x in row) + " |")

        else:

            for val in df[0]:
                if pd.notna(val):
                    md.append(str(val))

        md.append("")

    output_name = slugify(file.stem) + ".md"
    output_path = Path(OUTPUT_FOLDER) / output_name

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md))

    print("Created:", output_path)